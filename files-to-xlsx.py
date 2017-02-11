import openpyxl

import os

workbook = openpyxl.Workbook()
worksheet = workbook.active

files = [f for f in os.listdir(os.getcwd()) if os.path.isfile(f)]

for i in range(len(files)):
    #print(files[i])
    #print(type(files[i].encode('utf-8')))
    worksheet.cell(row=i+1, column=1, value=unicode(files[i],encoding='mbcs'))
    #worksheet.cell(row=i+1, column=1, value=files[i])

workbook.save('test.xlsx')
